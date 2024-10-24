from audioop import reverse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from django.conf import settings
import re
import os, shutil
import xlsxwriter
import pandas as pd
from datetime import datetime, timedelta

def clear_or_create_folder(bdir):
    if not os.path.exists(bdir):
        os.makedirs(bdir)
    else:
        for filename in os.listdir(bdir):
            file_path = os.path.join(bdir, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    
                    if  datetime.now().timestamp()- os.path.getmtime(file_path)>300:
                        os.unlink(file_path)

                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                print('Failed to delete %s. Reason: %s' % (file_path, e))

# def createExcel(name, data, headers,srt, datefields=[], namen="", userid = 0 ):  
    
#     help_text=["Старая" , "Новая", "ВПНД"]
#     if namen=="Список НД":        
#         headers.append( {'text': 'Признак', 'value': 'type_error'})    
#     font_size=11
#     fullpath = bdir.joinpath(name)
#     bdir = settings.BASE_DIR.joinpath('static').joinpath('downloadforuser').joinpath(str(userid))
#     clear_or_create_folder(bdir) 

#     wb = Workbook()
#     ws = wb.active
#     ws.cell(row=1, column=1).font = Font(bold = True)
#     ws.cell(row=1, column=1).value = namen    
#     ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

#     ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
#     cols_dict={}

#     for i, col in enumerate(headers):
#         ws.cell(row=2, column=i+1).font=Font(bold = True)
#         ws.cell(row=2, column=i+1).value = col['text'] 
#         putlengthtodict(ws.cell(row=2, column=i+1),cols_dict)

#     wb.save(fullpath)    
    # for indexR, row in enumerate(data):
    #     for datefield in datefields:           
    #         row[datefield] = row[datefield].strftime('%d.%m.%Y') if row[datefield] else ''
    #     for indexC, col in enumerate(headers):
    #         ws.cell(row=indexR+3, column=indexC+1).value = row[col['value']] if headers[indexC]['text']!="Признак" else help_text[int(row[col['value']])-1]
    #         putlengthtodict(ws.cell(row=indexR+3, column=indexC+1),cols_dict)
            
    # for [key,val] in cols_dict.items():         
    #     wid = val*font_size**(font_size*0.007)      
    #     ws.column_dimensions[key].width = wid if wid<100 else 100
    
    # FullRange = "A2:" + get_column_letter(ws.max_column) + str(ws.max_row)
    # ws.auto_filter.ref = FullRange
    # ws.freeze_panes = 'A3'
       
    
    
    # return str(fullpath)

    # columnssorter = {
    #     'MonitoringDataFBoneView':[],
    #     'JustifyDataFBoneView':[],
    #     'ListDataFBoneView':['date_unloading','region__tax_code','region__name', 'nocode','rocode','passport__code','ogrn','inn','fid','type_error']
    #     }






def createExcel(data, headers, namen="", userid = 0,types='ListDataFBoneView' ):
    nowtime=datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    
    headerrenamer={}
    for el in headers:
        headerrenamer[el['value']]=el['text']

    bdir = settings.BASE_DIR.joinpath('static').joinpath('downloadforuser').joinpath(str(userid))
    fullpath = bdir.joinpath(f'{namen.replace(" ","_")}.xlsx')
    clear_or_create_folder(bdir) 
    
    namen2=namen.replace("-",".").replace("_"," ")
    df = pd.DataFrame(data)
    # .reindex(columns=columnssorter[types])  
    if 'justify__methodologist__last_name' in df:
        
        df=df.drop(columns=['justify__methodologist__last_name', 'justify__methodologist__first_name','justify__methodologist__middle_name'])      
    if 'type_error' in df:
        terror = {1 : 'Старая', 2 : 'Новая', 3: 'ВПНД'}        
        df.replace({"type_error": terror},inplace=True)
    m=0
    shname = namen2
    if len(namen2)>=30:
        m = re.search("\d",namen2)
       
        shname = namen2[0: m.start()-1]
    df=df.rename(columns=headerrenamer, errors="ignore")
    writer = pd.ExcelWriter(fullpath, engine='xlsxwriter', date_format='dd.mm.yyyy')    
    df.to_excel(writer, sheet_name=shname,startrow=2,  header=False, index=False,freeze_panes=(2,0))
    workbook  = writer.book
  
    worksheet = writer.sheets[shname] 
    (max_row, max_col) = df.shape  
    header_format = {
    'text_wrap':'true',  
    'bold': 1,
    'border': 0,
    'align': 'center',
    'valign': 'vcenter',
    }
# Write the column headers with the defined format.
    for col_num, value in enumerate(df.columns.values):
        worksheet.write(1, col_num, value, workbook.add_format(header_format))   


    for i, width in enumerate(get_col_widths(df)):
        if width>100:
            width=100
        worksheet.set_column(i-1, i-1, width)
        

    worksheet.autofilter(1, 0, max_row, max_col-1)
    worksheet.set_row(0, 30)
    merge_format = workbook.add_format(header_format)
    
    worksheet.merge_range(0, 0, 0, max_col-1, namen2, merge_format)
    
    writer.save()
    fpn =  f'{settings.STATIC_URL}downloadforuser/{str(userid)}/{namen.replace(" ","_")}.xlsx'
    return str(fpn[1:])
    

def get_col_widths(dataframe):
    # First we find the maximum length of the index column   
    idx_max = max([len(str(s)) for s in dataframe.index.values] + [len(str(dataframe.index.name))])
    # Then, we concatenate this to the max of the lengths of column name and its values for each column, left to right
    return [idx_max] + [max([len(str(s)) for s in dataframe[col].values] + [len(col)+8]) for col in dataframe.columns]






def putlengthtodict(cell,cols_dict):
    letter = cell.column_letter
    len_cell = len(str(cell.value))
    len_cell_dict=0
    if letter in cols_dict:
        len_cell_dict=cols_dict[letter]
    if len_cell>len_cell_dict:
        cols_dict[letter]=len_cell
    if len_cell>100:
        cell.alignment = Alignment(wrapText=True)
